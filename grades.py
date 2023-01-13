from selenium import webdriver

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import time
import re

import warnings

def fxn():
    warnings.warn("deprecated", DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    fxn()

reshema_A = ["234302","234303","234304","234313","234326","234329","234493","234901","236025","236026","236201","236216","236268","236271","236272","236278","236299","236303","236304","236306","236309","236310","236313","236315","236318","236319","236321","236322","236323","236324","236328","236329","236330","236332","236333","236334","236336","236340","236341","236342","236345","236346","236347","236348","236349","236350","236351","236356","236357","236358","236359","236360","236361","236363","236366","236369","236370","236371","236372","236374","236376","236377","236378","236379","236381","236388","236422","236490","236491","236496","236499","236500","236501","236502","236503","236504","236506","236508","236509","236510","236512","236513","236515","236518","236520","236521","236522","236523","236524","236525","236526","236612","236613","236620","236621","236622","236623","236624","236625","236627","236628","236629","236630","236631","236632","236633","236634","236635","236637","236638","236640","236641","236643","236644","236645","236646","236647","236648","236649","236650","236651","236652","236653","236654","236655","236657","236658","236660","236661","236662","236663","236664","236667","236698","236700","236703","236712","236715","236716","236719","236729","236754","236755","236756","236757","236760","236777","236779","236780","202320","236781","236800","236811","236812","236813","236814","236815","236816","236817","236818","236819","236820","236821","236822","236823","236824","236825","236826","236827","236828","236829","236830","236831","236832","236833","236834","236835","236836","236837","236838","236860","236861","236862","236873","236874","236875","236901","236927","236990","236991","238125","238739","238790","238900","238901","238902"]
reshema_B = ["036044","044105","044127","044131","044137","044157","044167","044169","044202","046201","046206","046332","046880","048878","048921","086761","094222","094313","094314","094333","094334","094423","094591","096224","096250","096262","096326","096411","097317","104122","104135","104142","104157","104165","104174","104158","104177","104192","104221","104223","104279","104293","106378","104294","106383","114101","114246","115203","115204","114036","116217","116354","124120","124400","124503","124801","125801","134019","134020","134058","134082","134113","134128","134119","134142","214909"]
malagem = ["214119","214120","275112","324265","324267","324269","324273","324274","324258","324282","324284","324286","324292","324294","324297","324307","324314","324424","324433","324439","324441","324442","324445","324446","324527","324528","324541","324975","324992","325006","326000","326001","326002","326005","326006"]


options = Options()
options.headless = True
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)

print("Pick List of Courses you're interested in: ")
print("A for CS Reshima A")
print("B for CS Reshima B")
print("M for Malagem")
print("or 3 Digits that represent the faculty (e.g. 234,104..)")
x = ""
list = []
while x == "":
    x = input("Please input a valid Char:  ")
    if x == "A":
        list = reshema_A
    if x == "B":
        list = reshema_B
    if x == "M":
        malagem_url = "https://ugportal.technion.ac.il/%D7%9C%D7%99%D7%9E%D7%95%D7%93%D7%99-%D7%94%D7%A2%D7%A9%D7%A8%D7%94/"
        driver.get(malagem_url)
        body_text = driver.find_element(By.TAG_NAME,'body').text
        list = re.findall("[0-9]{6}",body_text)
    elif  x!="" :
        malagem_url = "https://michael-maltsev.github.io/technion-histograms/"
        driver.get(malagem_url)
        body_text = driver.find_element(By.TAG_NAME, 'body').text
        list = re.findall(x+"[0-9]{3}", body_text)





wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Number"
ws.cell(row=1, column=3).value = "Average"
ws.cell(row=1, column=4).value = "Median"
ws.cell(row=1, column=5).value = "Histograms Available"
for i in {1,2,3,4,5}:
    ws.cell(row=1, column=i).alignment = Alignment(horizontal='center',vertical='center')
    ws.cell(row=1, column=i).font = Font(size=15,bold=True)
    ws.column_dimensions[get_column_letter(i)].width = 50


for i in range(list.__len__()):
    url = 'https://michael-maltsev.github.io/technion-histograms/{}'.format(list[i])
    title_url = 'https://students.technion.ac.il/local/technionsearch/course/{}'.format(list[i])
    driver.get(url)
    time.sleep(2)
    tables = driver.find_elements(By.TAG_NAME,'table')
    count = 0
    total_avg = 0
    total_med = 0
    for table in tables:
        rows = table.find_elements(By.TAG_NAME, "tr")
        if rows.__len__() == 2:
            cols = table.find_elements(By.TAG_NAME, "td")
            if cols.__len__() == 7:
                if cols[5].text != '' and cols[5].text != ' ':
                    total_avg += float(cols[5].text)
                    total_med += float(cols[6].text)
                    count += 1
    if count>0:
        total_avg /= count
        total_med /= count
    driver.get(title_url)
    time.sleep(2)
    course_name = driver.find_element(By.TAG_NAME,'h1').text
    ws.cell(row=i+2,column=1).value = course_name
    ws.cell(row=i+2,column=1).alignment = Alignment(horizontal='center',vertical='center')
    ws.cell(row=i+2,column=2).value = list[i]
    ws.cell(row=i+2,column=2).alignment = Alignment(horizontal='center',vertical='center')
    ws.cell(row=i+2, column=3).value = total_avg
    ws.cell(row=i+2, column=3).alignment = Alignment(horizontal='center',vertical='center')
    ws.cell(row=i + 2, column=4).value = total_med
    ws.cell(row=i + 2, column=4).alignment = Alignment(horizontal='center',vertical='center')
    ws.cell(row=i + 2, column=5).value = (count > 0)
    print('Saved {}'.format(list[i]))
    time.sleep(2)
wb.save('{}_List.xlsx'.format(x))
